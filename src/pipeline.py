import pandas as pd
import logging
from datetime import datetime
from numbers import Number
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Dict, Optional, Tuple
from src.config import crear_carpeta_reporte

logger = logging.getLogger(__name__)

# Estilado para encabezados
FORMATO_ENCABEZADO = {
    "fill": PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid"),
    "font": Font(bold=True, color="FFFFFF", size=12),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "border": Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
}

# Configuración de formatos
CATEGORIAS_FORMATO = {
    "fecha": {
        "columna": [
            # Ingresos-GAIA
            "fecha_ingreso", "fecha_creacion", "fecha_ultimo_ingreso", 

            # EstadosCuenta-GAIA
            "Fecha_Contrato", "Fecha_Firma_Contrato", "Fecha_Proximo_pago", "fecha_promesa",

            # Ingresos-Condominios
            "FECHA_INGRESO", "FECHA_REGISTRO",

            # Ingresos-Condominios-BI
            "FechaPago",

            # EstadosCuenta-Condominios
            # "Fecha_Proximo_pago",
            "Fecha_ultimo_ingreso",
            
            # CarteraVencida-Condominios
            "FECHA_NACIMIENTO", "FECHAPAGO",
        ],
        "formato_excel": "DD/MM/YYYY",
    },
    # "fecha_hora": {
    #     "columna": [
    #         "timestamp", "fecha_hora", "datetime"
    #     ],
    #     "formato_excel": "DD/MM/YYYY HH:MM",
    # },
    "moneda": {
        "columna": [
            # Ingresos-GAIA
            "Cantidad", "Gastos_gestion",

            # EstadosCuenta-GAIA
            "Precio_venta", "Total_cobrado", "Enganche_pagado", "Total_por_cobrar", "Mensualidad", "Total_Requerido", "Cobrado", "Acumulado_Vencido", "Monto_ultimo_ingreso", "Acumado_Vencido",

            # Ingresos-Condominios
            "MONTO_PAGADO", "SALDO_PENDIENTE_POR_APLICAR", "MONTO_CUOTA", "MONTO_RESERVA", "MONTO_FONDO",

            # Ingresos-Condominios-BI
            "Monto", "MontoCuota", "MontoReserva", "MontoFondo", "FondosFuturos",

            # EstadosCuenta-Condominios
            # "Total_por_cobrar", "Total_cobrado",,m
            "Total_generado", "Saldo_vencido", "Saldo_pendiente_por_aplicar",
            
            # CarteraVencida-Condominios
            "TOTAL_PAGO", "SALDO_VENCIDO",
        ],
        "formato_excel": '"$"#,##0.00',
    },
    "numero": {
        "columna": [
            # Ingresos-GAIA
            "id_venta", "id_ingreso",

            # EstadosCuenta-GAIA
            # "id_venta",
            "dia_pago", "Meses_Financia", "Dias_Atrasado", "numero_pago",

            # Ingresos-Condominios
            "IDINGRESO", "IDCLIENTE", "NUMERO_CUENTA", "id_ingreso_dt",

            # Ingresos-Condominios-BI
            #"id_ingreso", "id_venta",

            # EstadosCuenta-Condominios
            "id_cliente", "Dias_atraso", "Dia_pago",
            
            # CarteraVencida-Condominios
            "IDCLIENTE", "DIA_VENCIDO",
        ],
        "formato_excel": '0',
    },
    # "numero_decimal": {
    #     "columna": [
    #         "metros", "unidades", "numero", "count", "cantidad_total",
    #     ],
    #     "formato_excel": '#,##0',
    # },
    "texto": {
        "columna": [
            # Ingresos-GAIA
            "id", "Marca", "Desarrollo", "Privada", "Etapa", "Unidad", "Folio", "Cliente", "STP", "Estatus", "forma_de_pago", "concepto", "flujo_concepto", "Banco", "Usuario_asignacion",

            # EstadosCuenta-GAIA
            # "id", "Marca", "Desarrollo", "Privada", "Etapa", "Unidad", "Cliente", "Estatus",
            "Copropietario", "Asesor", "Sucursal", "Tipo", "Equipo", "telefono_celular", "correo_electronico", "CuentaBeneficiarioReal",

            # Ingresos-Condominios
            # "STP",
            "DESARROLLO", "UNIDAD", "FOLIO", "CLIENTE", "BANCO", "FORMA_PAGO", "USUARIOS_REGISTRO", "STATUS", "SISTEMA",

            # Ingresos-Condominios-BI
            # "Marca", "Desarrollo", "Privada", "Etapa", "Unidad", "Cliente", "Banco",
            "folio", "Usuario", "cuentaBeneficiario", "FormaPago",

            # EstadosCuenta-Condominios
            # "Marca", "Desarrollo", "Unidad", "Etapa", "Cliente",
            "Id", "Correo", "Telefono", "Beneficiario_STP",
            
            # CarteraVencida-Condominios
            # "DESARROLLO", "UNIDAD", "CLIENTE", "SISTEMA",
            "CORREO", "TELEFONO", "nombre",
        ],
        "formato_excel": '@',
    },
}

class FormatoExcel:
    """Manejador de formatos para archivos Excel"""
    
    @staticmethod
    def detectar_categoria_columna(nombre_columna: str) -> Optional[str]:
        """Detecta la categoría de una columna basada en palabras clave"""
        if not nombre_columna:
            return None
        
        nombre = str(nombre_columna).strip()
        
        for categoria, config in CATEGORIAS_FORMATO.items():
            for keyword in config["columna"]:
                if keyword in nombre:
                    return categoria
        
        return None
    
    @staticmethod
    def obtener_formato_para_columna(nombre_columna: str) -> Optional[str]:
        """Obtiene el formato Excel para una columna basado en su nombre"""
        categoria = FormatoExcel.detectar_categoria_columna(nombre_columna)
        return CATEGORIAS_FORMATO.get(categoria, {}).get("formato_excel") if categoria else None
    
    @staticmethod
    def normalizar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Normaliza tipos de datos para que Excel aplique formatos correctamente"""
        df = df.copy()
        
        for col in df.columns:
            categoria = FormatoExcel.detectar_categoria_columna(col)
            
            if categoria == "fecha":
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    if hasattr(df[col], 'dt'):
                        df[col] = df[col].dt.tz_localize(None)
                except Exception:
                    logger.debug(f"No se pudo convertir columna {col} a fecha")
                    
            elif categoria == "moneda":
                try:
                    # Limpiar caracteres no numéricos como $, , (comas), espacios
                    if df[col].dtype == 'object':
                        # Remover símbolos de moneda, comas y espacios
                        df[col] = df[col].astype(str).str.replace(r'[^\d\.\-]', '', regex=True)
                        # Reemplazar múltiples puntos por uno solo (para decimales)
                        df[col] = df[col].str.replace(r'\.+', '.', regex=True)

                    # Convertir a numérico
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                    # Redondear a 2 decimales para moneda
                    df[col] = df[col].round(2)
                    
                except Exception as e:
                    logger.warning(f"No se pudo convertir columna {col} a moneda: {str(e)}")
                    
            elif categoria == "numero":
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype('Int64')

            # elif categoria == "numero_decimal":
            #     df[col] = pd.to_numeric(df[col], errors="coerce").round(2)
                
            elif categoria == "texto":
                df[col] = df[col].fillna('').astype(str).str.strip()
        
        return df
    
    @staticmethod
    def aplicar_formato_excel(archivo_excel: Path) -> bool:
        """Aplica formato profesional a un archivo Excel"""
        try:
            workbook = load_workbook(archivo_excel)
            worksheet = workbook.active
            
            # Aplicar formato a encabezados
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = FORMATO_ENCABEZADO["fill"]
                cell.font = FORMATO_ENCABEZADO["font"]
                cell.alignment = FORMATO_ENCABEZADO["alignment"]
                cell.border = FORMATO_ENCABEZADO["border"]
            
            # Ajustar ancho de columnas
            for column_cells in worksheet.columns:
                column_letter = column_cells[0].column_letter
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value),
                    default=0
                )
                adjusted_width = min(50, max(8, (max_length + 2) * 1.1))
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formatos específicos a datos
            columnas_formateadas = 0
            total_columnas = 0
            
            font_data = Font(size=12)
            alignment_data = Alignment(horizontal='left', vertical='center')

            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell.value:
                    total_columnas += 1
                    formato = FormatoExcel.obtener_formato_para_columna(header_cell.value)

                    if formato:
                        columnas_formateadas += 1
                        # Aplicar formato específico y estilo general
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value is not None:
                                # Formato numérico
                                if "$" in formato and isinstance(cell.value, (int, float)):
                                    cell.number_format = formato
                                    if isinstance(cell.value, float):
                                        cell.value = round(cell.value, 2)
                                else:
                                    cell.number_format = formato
                                # Estilo general
                                cell.font = font_data
                                cell.alignment = alignment_data
            
            workbook.save(archivo_excel)
            logger.info(f"Formato aplicado a {columnas_formateadas}/{total_columnas} columnas")
            return True
            
        except Exception as e:
            logger.error(f"Error aplicando formato Excel: {str(e)}")
            return False


def crear_excel_con_formato(df: pd.DataFrame, output_path: Path, nombre_consulta: str) -> bool:
    """
    Crea un archivo Excel con formato aplicado automáticamente.
    
    Args:
        df: DataFrame de pandas
        output_path: Ruta donde guardar el archivo
        nombre_consulta: Nombre de la consulta para nombrar la hoja
    
    Returns:
        bool: True si se creó correctamente
    """
    try:
        # Normalizar y guardar DataFrame
        df = FormatoExcel.normalizar_dataframe(df)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            nombre_hoja = nombre_consulta[:31]  # Excel limita a 31 caracteres
            df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        
        # Aplicar formato
        return FormatoExcel.aplicar_formato_excel(output_path)
        
    except Exception as e:
        logger.error(f"Error creando Excel con formato: {str(e)}")
        return False


def procesar_bigquery_dataframe(dataframes_dict: Dict[str, pd.DataFrame]) -> Tuple[Dict[str, Path], Path]:
    """
    Procesar múltiples DataFrames y guardar cada uno en archivo Excel con formato.
    
    Returns:
        Tuple[Dict[str, Path], Path]: Resultados y carpeta de reporte
    """
    carpeta_reporte = crear_carpeta_reporte()
    logger.info(f"Carpeta de reporte: {carpeta_reporte}")
    
    resultados = {}
    
    for nombre_consulta, df in dataframes_dict.items():
        if df.empty:
            logger.warning(f"DataFrame vacío para {nombre_consulta}")
            continue
        
        try:
            nombre_archivo = f"{nombre_consulta}.xlsx"
            output_path = carpeta_reporte / nombre_archivo
            
            logger.info(f"Procesando: {nombre_consulta}")
            
            if crear_excel_con_formato(df, output_path, nombre_consulta):
                resultados[nombre_consulta] = output_path
                logger.info(f"✓ Guardado con formato: {output_path.name}")
            else:
                logger.warning(f"⚠ Guardado sin formato: {output_path.name}")
                resultados[nombre_consulta] = output_path
                
        except Exception as e:
            logger.error(f"Error procesando {nombre_consulta}: {str(e)}")
    
    return resultados, carpeta_reporte